#!/bin/python3

####################################################################################
# This Script is developed to run Image scan in AWS ECR and Upload the result in local
# in a excel worksheet
# ####################################################################################

import subprocess
import yaml
import json
import pandas as pd
from openpyxl import load_workbook ,Workbook

REGION = "us-west-2"

REGISTRY_ID = "xxxxxxx564"

PROFILE = "my-aws"
INPUT_FILE = "input.yaml"

RELEASE = ""
RELEASE_FILE = ""
IMG_NAME = ""
REPO = ""

TAG = ""


# This function will take info from input.yaml and return  a nested list of that information

def read_input_file(file_name):
    RELEASE_NAME = []
    IMAGE_NAME = []
    REPO_NAME = []
    IMAGE_TAG = []
    with open(file_name, "r") as f:
        data = yaml.safe_load(f)  
    RELEASE_NAME += [f"release-{data['release']['version']}"]
    for component in data['images']:
        IMAGE_NAME += [component["image"]]
        REPO_NAME += [component["repository"]]
        IMAGE_TAG += [component["tag"]]
    ALL_IMG_INFO = [RELEASE_NAME, IMAGE_NAME,REPO_NAME,IMAGE_TAG] 
    return ALL_IMG_INFO


# This function will create json file of vulnerabilities scanning for given images in input file
def image_scan_info():
    try:
        get_cmd = f"aws ecr describe-image-scan-findings --registry-id {REGISTRY_ID} --repository-name {REPO} --image-id imageTag={TAG} --region {REGION} --query imageScanFindings --profile {PROFILE} --output json"
        get_scan = subprocess.run(get_cmd, shell=True, capture_output=True, text=True)
        if get_scan.returncode != 0:
            print(f"\nError running command for Image  {IMG_NAME}:{TAG}. Return code: {get_scan.returncode}\nPlease check the image info or  manually start the ECR scan for the image tag\n")
            print(get_scan.stderr)
        data = json.loads(get_scan.stdout)
        final_data =  processDataToTable(data)
        append_to_excel(final_data, RELEASE_FILE, RELEASE)
    except Exception as e:
        print(f"\nException occurred for Image  {IMG_NAME}:{TAG}: {str(e)}")
        print(f"\nUnsupported Image Error from ECR. Please manually put the scan results from docker hub\n")
        
        
# This function will create a excel file
def createExcelFile(excel_file_path, sheet_name):
    from openpyxl import Workbook, load_workbook
    import os
    if os.path.isfile(excel_file_path):
        # If the Excel file exists, load it
        try:
            workbook = load_workbook(excel_file_path)
            if sheet_name in workbook.sheetnames:
                print(f"The sheet '{sheet_name}' exists in the Excel file.")
            else:
                # If the sheet doesn't exist, create it
                worksheet = workbook.create_sheet(sheet_name)
                workbook.save(excel_file_path)
                print(f"The sheet '{sheet_name}' has been created in the Excel file.")
        except Exception as e:
            print(f"Error while loading the Excel file: {e}")
    else:
        # If the Excel file doesn't exist, create it and the sheet
        try:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = sheet_name
            workbook.save(excel_file_path)
            print(f"The Excel file '{excel_file_path}' has been created with the sheet '{sheet_name}'.")
        except Exception as e:
            print(f"Error while creating the Excel file: {e}")
            
# This function will append dataframe data to excel sheet            
def append_to_excel(final_data, excel_file_path, sheet_name):
    import pandas as pd
    try:
        # Load the existing Excel file, or create a new one if it doesn't exist
        with pd.ExcelWriter(excel_file_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            # Try to load the existing sheet, or create a new sheet if it doesn't exist
            try:
                existing_data = pd.read_excel(excel_file_path, sheet_name=sheet_name)
                new_data = pd.concat([existing_data, final_data], ignore_index=True)
                # Write the updated data to the same sheet
                new_data.to_excel(writer, index=False, sheet_name=sheet_name)
            except pd.errors.EmptyDataError:
                # If the sheet doesn't exist, create a new sheet and write the data
                final_data.to_excel(writer, index=False, sheet_name=sheet_name, engine='xlsxwriter')
        print("Scan Data appended successfully.")
    except Exception as e:
        print(f"Error while appending data to Excel file: {e}")
# image_scan_info()

def processDataToTable(data):
    import pandas as pd
    df = pd.json_normalize(data)
    columns = [
      'findingSeverityCounts.HIGH',
      'findingSeverityCounts.MEDIUM',
      'findingSeverityCounts.LOW',
      'findingSeverityCounts.UNDEFINED',
      'findingSeverityCounts.INFORMATIONAL',
      'findingSeverityCounts.CRITICAL'
    ]
    for col in columns:
        if col not in df.columns:
            df[col] = 0
    vulnerability_codes = None
    if  df["findings"].any():
        vulnerability_codes = pd.DataFrame(df["findings"][0])['name'].tolist()
    
    df.pop('findings')
    df.pop('findingSeverityCounts.UNDEFINED')
    df.pop('findingSeverityCounts.INFORMATIONAL')
    df.pop('imageScanCompletedAt')
    df.pop('vulnerabilitySourceUpdatedAt')
    order_column = ['Images', 'CRITICAL', 'HIGH', 'MEDIUM', 'LOW', 'vulnerability_codes']
    df['Images'] = "{}:{}".format(REPO, TAG)
    df.rename(
    columns = {
    'findingSeverityCounts.CRITICAL': 'CRITICAL',
    'findingSeverityCounts.HIGH':'HIGH', 
    'findingSeverityCounts.MEDIUM':'MEDIUM',
    'findingSeverityCounts.LOW': 'LOW'
    }, inplace = True)
    df['vulnerability_codes'] = [vulnerability_codes]
    return df[order_column]

      
def get_main():
    global RELEASE, RELEASE_FILE, IMG_NAME, REPO, TAG
    
    RELEASE = read_input_file(INPUT_FILE)[0][0]
    RELEASE_FILE = f"{RELEASE}.xlsx"
    createExcelFile(RELEASE_FILE, RELEASE)
    for i in range(len(read_input_file(INPUT_FILE)[1])):
        IMG_NAME = read_input_file(INPUT_FILE)[1][i]
        REPO = f"{read_input_file(INPUT_FILE)[2][i]}/{IMG_NAME}"
        TAG = read_input_file(INPUT_FILE)[3][i]

        image_scan_info()
    print("\nALL SCAN DATA SUCCESSFULLY APPENED, FILE IS READY !!!!")   
  
get_main()